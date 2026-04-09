"""
exporter.py
-----------
Exports extracted data records to CSV or Excel files.
"""

import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from rich.console import Console

console = Console()


def export_data(
    records: list[dict],
    fields: list[str],
    output_dir: str,
    output_file: str,
    output_format: str = "csv"
) -> str:
    """
    Exports extracted records to CSV or Excel.

    Args:
        records:       List of extracted data dicts
        fields:        Ordered list of field names (defines column order)
        output_dir:    Directory to save the output file
        output_file:   Base filename (without extension)
        output_format: "csv" or "excel"

    Returns:
        Full path to the saved output file
    """
    if not records:
        console.print("[yellow]⚠️  No records to export.[/yellow]")
        return ""

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # Build column list — user fields first, then source file column
    columns = fields + ["_source_file"]

    df = pd.DataFrame(records, columns=columns)

    # Clean up: strip whitespace from string fields
    for col in fields:
        if df[col].dtype == object:
            df[col] = df[col].str.strip() if hasattr(df[col], 'str') else df[col]

    # Add timestamp to filename to avoid overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{output_file}_{timestamp}"

    if output_format.lower() == "excel":
        file_path = os.path.join(output_dir, f"{base_name}.xlsx")
        _export_excel(df, file_path, fields)
    else:
        file_path = os.path.join(output_dir, f"{base_name}.csv")
        _export_csv(df, file_path)

    console.print(f"\n[bold green]💾 File saved:[/bold green] {file_path}")
    console.print(f"[dim]   Rows: {len(df)} | Columns: {len(fields)}[/dim]")

    return file_path


def _export_csv(df: pd.DataFrame, file_path: str):
    """Saves DataFrame as a UTF-8 CSV file."""
    df.to_csv(file_path, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compatibility


def _export_excel(df: pd.DataFrame, file_path: str, fields: list[str]):
    """Saves DataFrame as a formatted Excel file."""
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extracted Data")

        workbook = writer.book
        worksheet = writer.sheets["Extracted Data"]

        # Auto-fit column widths
        for col_idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 4
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = min(max_length, 50)

        # Style the header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    console.print(f"[dim]   Excel formatting applied.[/dim]")


def print_preview(records: list[dict], fields: list[str], max_rows: int = 5):
    """Prints a preview table of extracted data in the terminal."""
    if not records:
        return

    console.print(f"\n[bold]📋 Preview (first {min(max_rows, len(records))} records):[/bold]")

    df = pd.DataFrame(records[:max_rows], columns=fields + ["_source_file"])
    console.print(df.to_string(index=False))
    console.print()
